#openpyxl写入数据
#封装一个写入excel数据的方法
# coding:utf-8
from openpyxl import load_workbook
import openpyxl
import sys
sys.path.append("../..")
from config.config import *

#openpyxl写入数据
#封装一个写入excel数据的方法
# coding:utf-8
from openpyxl import load_workbook
import openpyxl
def copy_excel(excelpath1, excelpath2):
    '''复制excek，把excelpath1数据复制到excelpath2'''
    wb2 = openpyxl.Workbook()
    wb2.create_sheet("sheet1",1)
    wb2.save(excelpath2)
    # 读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    # 循环遍历所有sheet
    for i in range(len(sheets1)):
        sheet1 = wb1[sheets1[i]]
        sheet2=sheet1
        ws = wb2[sheets2[i]]

        ws.title = str(sheet2)  # 修改名为Sheet1工作表名称
        max_row = sheet1.max_row  # 最大行数
        max_column = sheet1.max_column  # 最大列数
        for m in list(range(1, max_row + 1)):
            for n in list(range(97, 97 + max_column)):  # chr(97)='a'
                n = chr(n)  # ASCII字符
                i = '%s%d' % (n, m)  # 单元格编号
                cell1 = sheet1[i].value# 获取data单元格数据
                sheet2[i].value = cell1  # 赋值到test单元格
        wb2.save(excelpath2)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()
def wirte_result(result, filename):
    # 返回结果的行数row_nub
    row_nub = result['rowNum']
    # 写入statuscode
    wt = Write_excel(filename)
    wt.write(row_nub, 10, result['statuscode'])       # 写入返回状态码statuscode,第8列
    wt.write(row_nub, 11, result['times'])            # 耗时
    wt.write(row_nub, 12, result['error'])            # 状态码非200时的返回信息
    wt.write(row_nub, 13, result['result'])           # 测试结果 pass 还是fail
    wt.write(row_nub, 14, result['msg'])
    wt.write(row_nub, 15, result['retInfo'])  # 抛异常
class Write_excel(object):
     '''修改excel数据'''
     def __init__(self, filename):
         self.filename = filename
         self.wb = load_workbook(self.filename)
         self.ws = self.wb.active  # 激活sheet

     def write(self, row_n, col_n, value):
         '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
         self.ws.cell(row_n, col_n).value = value
         self.wb.save(self.filename)
         self.wb.close()

if __name__ == "__main__":
     copy_excel(data_file, report_excel)
     #wt = Write_excel(report_excel)